from openpyxl import load_workbook
from typing import Optional


def read_from_xlsx(
    xlsx_path: str,
    col_key_map: dict,
    sheet_name: Optional[str] = None
    ) -> list[dict]:
    """
    从xlsx表格中自定义读取指定列的数据，并返回一个字典列表。
    
    Args:
        xlsx_path: xlsx文件路径
        col_key_map: 列名与键的映射关系，例如 {"A": "X", "B": "Y"}
        sheet_name: 表格名称，默认为None，表示使用第一个表格
    
    Returns:
        list[dict]: 字典列表，例如 [{"X": value1, "Y": value2}, ...]
    """
    # 加载工作簿
    workbook = load_workbook(xlsx_path, data_only=True)
    
    # 选择工作表
    if sheet_name is None:
        worksheet = workbook.active
    else:
        worksheet = workbook[sheet_name]
    
    # 获取要查找的列名列表
    col_names = list(col_key_map.keys())
    
    # 找到包含所有列名的行
    header_row = None
    col_indices = {}  # 存储列名对应的列索引（从1开始）
    
    for row_idx, row in enumerate(worksheet.iter_rows(values_only=False), start=1):
        # 检查当前行是否包含所有列名
        found_cols = {}
        for cell in row:
            if cell.value in col_names:
                found_cols[cell.value] = cell.column
                col_indices[cell.value] = cell.column
        
        # 如果找到了所有列名，记录这一行
        if len(found_cols) == len(col_names):
            header_row = row_idx
            break
    
    if header_row is None:
        workbook.close()
        raise ValueError(f"未找到包含所有列名 {col_names} 的行")
    
    # 从下一行开始读取数据
    result = []
    for row_idx in range(header_row + 1, worksheet.max_row + 1):
        row_data = {}
        has_data = False
        
        for col_name, key_name in col_key_map.items():
            col_idx = col_indices[col_name]
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            
            # 如果单元格有值，记录到字典中
            if cell_value is not None:
                row_data[key_name] = cell_value
                has_data = True
        
        # 只有当至少有一个非空值时，才添加到结果中
        if has_data:
            result.append(row_data)
    
    workbook.close()
    return result